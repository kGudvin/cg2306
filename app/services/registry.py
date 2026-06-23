from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import RegistryProduct


@dataclass
class RegistryImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def display_name_for_registry_product(name: str, registry_number: str) -> str:
    return f"{name}\nРеестровый номер: {registry_number}"


def normalize_registry_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        value = format(value, "f")
    elif isinstance(value, float) and value.is_integer():
        value = str(int(value))
    else:
        value = str(value)
    return value.strip()


def import_registry_products(db: Session, file_bytes: bytes, source_file_name: str) -> RegistryImportStats:
    stats = RegistryImportStats()
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        products_by_number = {
            product.registry_number: product
            for product in db.scalars(select(RegistryProduct)).all()
        }
        sheet = workbook.worksheets[0]
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            name = normalize_registry_value(row[0] if len(row) > 0 else None)
            registry_number = normalize_registry_value(row[3] if len(row) > 3 else None)
            if not name or not registry_number or not registry_number.isdigit():
                stats.skipped += 1
                continue

            product = products_by_number.get(registry_number)
            if product is None:
                product = RegistryProduct(
                    registry_number=registry_number,
                    name=name,
                    source_file_name=source_file_name,
                )
                db.add(product)
                products_by_number[registry_number] = product
                stats.created += 1
            else:
                if product.name != name or product.source_file_name != source_file_name:
                    product.name = name
                    product.source_file_name = source_file_name
                    product.updated_at = datetime.now()
                    stats.updated += 1
                else:
                    stats.skipped += 1
        db.commit()
    except Exception as exc:
        db.rollback()
        stats.errors.append(str(exc))
    finally:
        workbook.close()
    return stats
