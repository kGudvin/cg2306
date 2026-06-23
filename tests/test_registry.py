import unittest
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import RegistryProduct
from app.services.registry import display_name_for_registry_product, import_registry_products


def make_workbook_bytes(rows: list[tuple[str | None, str | None]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for index, (name, registry_number) in enumerate(rows, start=1):
        sheet.cell(index, 1).value = name
        sheet.cell(index, 4).value = registry_number
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class RegistryImportTest(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(bind=engine)
        self.session = sessionmaker(bind=engine)()

    def tearDown(self):
        self.session.close()

    def test_import_registry_products_upserts_and_skips_headers(self):
        data = make_workbook_bytes(
            [
                ("Реестр российской промышленной продукции", "реестр номер"),
                (None, None),
                ("Клавиатура компьютерная проводная БЕШТАУ КЛ104РУ, БЕРТ.467219.003-02", "10715802"),
            ]
        )

        first = import_registry_products(self.session, data, "registry.xlsx")
        second = import_registry_products(self.session, data, "registry.xlsx")
        product = self.session.scalar(select(RegistryProduct).where(RegistryProduct.registry_number == "10715802"))

        self.assertEqual(first.created, 1)
        self.assertEqual(first.skipped, 2)
        self.assertEqual(second.created, 0)
        self.assertEqual(second.updated, 0)
        self.assertEqual(second.skipped, 3)
        self.assertIsNotNone(product)
        self.assertEqual(product.name, "Клавиатура компьютерная проводная БЕШТАУ КЛ104РУ, БЕРТ.467219.003-02")
        self.assertEqual(
            display_name_for_registry_product(product.name, product.registry_number),
            "Клавиатура компьютерная проводная БЕШТАУ КЛ104РУ, БЕРТ.467219.003-02\nРеестровый номер: 10715802",
        )


if __name__ == "__main__":
    unittest.main()
